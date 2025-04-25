import pandas as pd
import numpy as np
import weightipy as wp
import math
from collections import defaultdict

from . import z_test


def create_table(
        df,
        meta,
        banner=None,
        ttl_var='ttl',
        wt_var=None,
        empty_incl=True,
        drop_mp_zeros=True,
        calc_r_cl=True,
        client_sublabels=False,
        calc_mean=True,
        cl=None
):


    def extract_aggregates(recode_dict):
        """
        Extracts aggregates from a dictionary.

        Parameters:
        - recode_dict (dict): A dictionary containing key-value pairs.

        Returns:
        - renamed_dict (dict): A new dictionary with renamed keys.

        Description:
        This function takes a dictionary, `recode_dict`, and extracts aggregates from it. It finds values that are not unique, counts the occurrences of each value, and creates a new dictionary where the keys are the values and the values are lists of keys that have that value. It then finds the top and bottom keys based on the minimum and maximum values in the dictionary, and renames them as 'Top-<agg_len>' and 'Bottom-<agg_len>' respectively, where `<agg_len>` is the length of the aggregate. Finally, it creates a new dictionary with the renamed keys and returns it.

        Note:
        - The function assumes that the values in the `recode_dict` are hashable.

        Example Usage:
        ```
        recode_dict = {1: 'top-2', 2:'top-2',3: 'bottom-2', 4:'bottom-2'}
        renamed_dict = extract_aggregates(recode_dict)
        print(renamed_dict)
        # Output: {'Top-2': [1, 2], 'Bottom-2': [3,4]}
        ```
        """
        # Find values that are not unique
        values_count = {}
        for key, value in recode_dict.items():
            values_count[value] = values_count.get(value, 0) + 1

        # Extract pairs where the value is not unique
        non_unique_dict = {key: value for key, value in recode_dict.items() if values_count[value] > 1}

        # Aggregate recode values to format {key: list of enclosed values}
        aggregate_dict = {}
        for key, value in non_unique_dict.items():
            if value not in aggregate_dict:
                aggregate_dict[value] = [key]
            else:
                aggregate_dict[value].append(key)

        # Find top and bottom keys
        top_key = min(aggregate_dict)
        bot_key = max(aggregate_dict)

        # Find len of aggregate
        agg_len = len(aggregate_dict[top_key])

        # Rename to top-bottom
        key_mapping = {top_key: f'Top-{agg_len}', bot_key: f'Bottom-{agg_len}'}

        # Create a new dictionary with renamed keys
        renamed_dict = {key_mapping.get(key, key): value for key, value in aggregate_dict.items()}

        return renamed_dict


    def fill_empty_dimension(table, val_labels, axis):
        """
        Fill dimension in a table with missing values.
        Parameters:
            - table: The table to fill empty dimension in.
            - val_labels: A dictionary containing the full value-labels dictionary.
            - axis: The axis along which to fill the empty dimension.
                    If axis is 0, the operation is performed on the rows.
                    If axis is 1, the operation is performed on the columns.
        Returns:
            - The table with the empty dimension filled.
        """
        exist_codes = table.index if axis == 0 else table.columns
        missing_codes = [code
                         for code in val_labels.keys()
                         if code != 0 and code not in exist_codes]

        if axis == 0:
            table = table.reindex(table.index.union(missing_codes), fill_value=0)
        else:
            table = table.reindex(columns=table.columns.union(missing_codes), fill_value=0)

        table.sort_index(axis=axis, inplace=True)
        return table


    def round_table(table_col, q_type):
        """
        A function to round the values in a table column and normalize to 1 if it is necessary.
        Parameters:
        - table_col: The column of the table to be rounded and processed. It should be a numerical array.
        - q_type: The type of question. If it is not 'multipunch', the column values will be normalized to 1.
        Returns:
        - table_col: The rounded and processed column of the table. It will have the same shape as the input column.
        """
        table_col = np.round(table_col, 2)

        # If not multipunch, normalize to 1
        if q_type != 'multipunch':

            # Find difference of 1
            diff = 1 - table_col.sum()
            # If column is empty or already fit to 1, skip it
            if diff == 0 or diff == 1:
                return table_col

            # Find the last non-zero row to apply math
            last_r = np.where(table_col != 0)[0][-1]
            # Depending on diff sign, apply math
            # If diff is negative, subtract diff from last row, update row if it falls to zero
            if diff < 0:
                # Anyway we need absolute to not going wild with subtraction
                diff = abs(diff)
                while diff > 0:
                    last_r_value = table_col.iloc[last_r]
                    if diff > last_r_value:
                        # If last row value is not enough, subtract all and move on
                        diff -= last_r_value
                        table_col.iloc[last_r] = 0
                        last_r -= 1
                    else:
                        # Or just subtract
                        table_col.iloc[last_r] -= diff
                        diff = 0
            else:
                # If positive, just add to last row
                table_col.iloc[last_r] += diff
        return table_col


    def calculate_mean_parameters(x):
        """
        Calculate the mean parameters.

        Parameters:
            x (pandas.DataFrame): The input data frame with variable and weight column.

        Returns:
            pd.Series: A series containing the calculated mean parameters:
                - 'Среднее': The weighted mean.
                - 'Среднеквадратичное отклонение': The weighted root mean square deviation.
                - 'Валидные': The sum of the weights.
        """
        wt_base = x[wt_var].sum()
        base_condition = wt_base > 1

        # To avoid excess computation, use if actual base exists
        if base_condition:
            wt_mean = np.average(x.iloc[:,0], weights=x[wt_var])
            wt_sq_deviations = np.average((x.iloc[:,0] - wt_mean) ** 2, weights=x[wt_var])
            if wt_sq_deviations != 0:
                wt_rmsd = np.sqrt(wt_sq_deviations * wt_base / (wt_base - 1))
            else:
                wt_rmsd = 0
        else:
            wt_mean = 0
            wt_rmsd = 0

        return pd.Series({
            'Среднее': wt_mean,
            'Среднеквадратичное отклонение': wt_rmsd,
            'Валидные': wt_base
        })


    def concat_row(table):
        """
        Concatenates the values of a dictionary of pandas DataFrames horizontally.

        Args:
            table (dict): A dictionary of pandas DataFrames.

        Returns:
            pandas.DataFrame: A new DataFrame obtained by concatenating the DataFrames
            horizontally.

        Example:
            >>> table = {'A': pd.DataFrame({'col1': [1, 2, 3], 'col2': [4, 5, 6]}),
            ...          'B': pd.DataFrame({'col1': [7, 8, 9], 'col2': [10, 11, 12]})}
            >>> concat_row(table)
               A        B
            col1 col2 col1 col2
            0    1    4    7    10
            1    2    5    8    11
            2    3    6    9    12
        """
        return pd.concat(list(table.values()), axis=1, keys=list(table.keys()))


    def set_row_labels(table, meta, row_var, sub_label):
        """
        Set the row labels of a table based on the given metadata.

        Parameters:
            table (pd.DataFrame): The table whose row labels will be set.
            meta (pd.DataFrame): The metadata that contains the necessary information.
            row_var (str): The variable name to be used for the row labels.
            sub_label (bool): A flag indicating whether to include a sub-label.

        Returns:
            pd.DataFrame: The table with the updated row labels.
        """
        table = table.rename(index=meta.var_value_labels[row_var])
        q_label = f"({str(row_var)[:5] if meta.q_type[row_var] == 'multipunch' else str(row_var)}) " if sub_label else ""
        table.index = pd.MultiIndex.from_tuples(
            [(f"{q_label}{meta.var_names_to_labels[row_var]}", value) for value in table.index], names=[None, None])
        return table


    def normalize_to_total(var_base, round_ttl_base, orig_ttl_base):
        """
        Normalize the given variable base to a total base.

        Parameters:
        - var_base: A pandas DataFrame representing the variable base.
        - ttl_value: The desired total value to normalize the variable base to.
        - orig_ttl_value: The original total value of the variable base.

        Returns:
        - var_base: The normalized variable base as a pandas DataFrame.
        """
        orig_sum = var_base.iloc[:, -1].sum()  # Find the base sum of the given question
        var_base = var_base.round(0)  # Round each base to the nearest integer
        round_sum = var_base.iloc[:, -1].sum()  # Find the new sum of the given question rounded bases

        # If the question is not filtered
        if orig_sum == orig_ttl_base:
            # Find normalizing coefficient
            coeff = round_ttl_base / round_sum

            # Normalize and round once again
            var_base.iloc[:, -1] = var_base.iloc[:, -1] * coeff
            var_base.iloc[:, -1] = var_base.iloc[:, -1].round()

            # Find delta of normalization and rounding
            delta = var_base.iloc[:, -1].sum() - round_ttl_base

            # Apply delta to the last non-zero value of variable
            if delta < 0:
                i = -1
                while var_base.iloc[i, -1].sum() == 0:
                    i -= 1
                var_base.iloc[i, -1] += abs(delta)
            elif delta > 0:
                i = -1
                while var_base.iloc[i, -1].sum() == 0:
                    i -= 1
                var_base.iloc[i, -1] -= abs(delta)
        return var_base


    # Setup weights and total variables
    if wt_var is None:
        wt_var = 'wt_default'
        df[wt_var] = 1

    if ttl_var not in df.columns:
        df[ttl_var] = 1

    if banner is None:
        banner = [ttl_var]

    # Initialize table templates
    pivot_tables = []
    round_tables = {}
    client_tables = []
    mean_tables = []
    abs_tables = []
    sig_tables = []
    sig_mean_tables = []

    # Initialize base series for client tables
    client_base = pd.Series([0])

    # Loop through all variables and all banner variables, generating pivot tables
    for row_i, row_var in enumerate(df.columns):
        # print(row_var)

        # Don't use variable if it is not single or multiple punch
        q_type_condition = meta.q_type[row_var] not in ['singlepunch', 'multipunch']
        # Don't use if empty include is set to False and variable is all nan
        empty_condition = df[row_var].isna().all()
        # If there is breaker inside the MP or gris block it is possible to duplicate columns
        if not isinstance(empty_condition, np.bool_):
            raise AttributeError(
            f'Проверьте переменную {df.columns[row_i - 1]} - рекоды в базе должны идти после всех переменных блока'
            )

        if q_type_condition or empty_condition:
            continue

        # Analyze the variable type
        is_mean = row_var in meta.scale_vars
        is_aggr = row_var in meta.scale_recode_for_aggregate and not row_var in meta.awareness_vars
        is_awareness = row_var in meta.awareness_vars
        is_orig = row_var in meta.original_vars
        rv_type = meta.q_type[row_var]
        # If base row should be applied
        is_last_var = row_i == len(df.columns) - 1
        is_last_mp_block = meta.var_names_to_labels[row_var] != meta.var_names_to_labels[
            df.columns[row_i + 1]] if rv_type == 'multipunch' and not is_last_var else False
        is_base = any([is_last_var, is_last_mp_block, rv_type == 'singlepunch'])

        # Initialize row variable tables
        r_pivot_table = {}
        r_round_table = {}
        r_cl_table = {}
        r_mean_table = {}
        r_abs_table = {}

        # Iterate through each banner variable, creating table and storing into row table
        for col_var in banner:
            # If banner var is empty, skip it
            banner_empty_condition = df[col_var].isna().all()
            if banner_empty_condition:
                continue

            # Calculate pivot table, fill nan with 0
            col_table = df.groupby([row_var, col_var])[wt_var].sum().unstack().fillna(0)

            # Drop zeroes from column
            if 0 in col_table.columns:
                col_table.drop(0, axis=1, inplace=True)

            # Fill rows with missed values for total and fill columns for the rest of banner
            if empty_incl:
                if col_var == ttl_var:
                    col_table = fill_empty_dimension(
                        table=col_table,
                        val_labels=meta.var_value_labels[row_var],
                        axis=0
                    )
                else:
                    col_table = fill_empty_dimension(
                        table=col_table,
                        val_labels=meta.var_value_labels[col_var],
                        axis=1
                    )

            # Save abs table
            abs_table = col_table.copy()

            # Calculate base for table
            base = col_table.sum()

            # Convert to percentage
            col_table = col_table / base

            # Drop zeroes from rows for multipunch
            if all([drop_mp_zeros, rv_type == 'multipunch', 0 in col_table.index]):
                col_table.drop(0, axis=0, inplace=True)
                abs_table.drop(0, axis=0, inplace=True)

            # Calculate mean tables
            if all([calc_mean, is_mean, is_orig]):
                temp_mean_series = df.loc[df[col_var] != 0, [row_var, col_var, wt_var]]
                if row_var != col_var:
                    temp_mean_series[row_var] = temp_mean_series[row_var].replace(meta.scale_recode_for_mean[row_var])
                # Handler to cross the variable to itself
                else:
                    temp_mean_series.columns = ['dummy_row', col_var, wt_var]
                    temp_mean_series['dummy_row'] = temp_mean_series['dummy_row'].replace(
                        meta.scale_recode_for_mean[row_var]
                    )

                temp_mean_series.dropna(axis=0, inplace=True)
                if temp_mean_series.empty:
                    c_mean_table = pd.DataFrame(
                        data={1: [np.nan] * 3},
                        index=['Среднее', 'Среднеквадратичное отклонение', 'Валидные'],
                    ).rename_axis(mapper=col_var, axis='columns')
                else:
                    c_mean_table = temp_mean_series.groupby(col_var, dropna=False).apply(calculate_mean_parameters).T

                # Fill empty if necessary
                if col_var != ttl_var:
                    c_mean_table = fill_empty_dimension(
                        table=c_mean_table,
                        val_labels=meta.var_value_labels[col_var],
                        axis=1
                    )

                # Append to the row table
                r_mean_table[col_var] = c_mean_table

            # Rounded and client tables with top-bottom
            if calc_r_cl:
                # Calculate rounded table
                # Round to 2 digits after fot for original variables, else aggregate through sum
                if is_orig:
                    c_round_table = col_table.apply(round_table, axis=0, args=(meta.q_type[row_var],))
                else:
                    # Extract aggregate codes for original variable
                    orig_var = row_var[4:]

                    # Find the original variable rounded row table
                    orig_table = round_tables[orig_var]
                    # Extract current column table
                    c_round_table = orig_table.loc[:, col_var]
                    c_round_table = c_round_table.reset_index(drop=True)
                    c_round_table.index = range(1, len(c_round_table) + 1)

                    # Reaggregate table
                    k = c_round_table.index
                    v = list(meta.scale_recode_for_aggregate[orig_var].values())
                    if 0 in v:
                        v.remove(0)
                    mapper = dict(zip(k, v))
                    c_round_table = c_round_table.groupby(c_round_table.index.map(mapper)).sum()

                # Process the client table
                if is_orig:
                    c_cl_table = c_round_table.iloc[:-1].copy()
                    cl_table_dk = c_round_table.iloc[-1:].copy()

                    dk_check = True

                    if is_aggr:
                        # Extract aggregate codes and add to the end
                        agg_helper = extract_aggregates(meta.scale_recode_for_aggregate[row_var])

                        # Check if there is DK considered in recode. If so, cocnat table back
                        if not cl_table_dk.empty:
                            if any([cl_table_dk.index[0] in agg_helper[agg] for agg in agg_helper]):
                                c_cl_table = pd.concat([c_cl_table, cl_table_dk], axis=0)
                                dk_check = False

                        for agg in agg_helper:
                            exist = [i for i in agg_helper[agg] if i in c_cl_table.index]
                            c_cl_table.loc[agg] = c_cl_table.loc[exist].sum()

                    elif is_awareness:
                        # Custom handler, fix it later
                        dk_check = False

                        c_cl_table = pd.concat([
                            c_cl_table,
                            cl_table_dk
                        ], axis=0)

                        # If awareness question filtrates column var, codes might be missing
                        c_cl_table = fill_empty_dimension(
                            table=c_cl_table,
                            val_labels=meta.var_value_labels[row_var],
                            axis=0
                        )

                        c_cl_table.loc['Знаю'] = c_cl_table.iloc[:2].sum()
                        c_cl_table.loc['Не знаю'] = c_cl_table.iloc[2]


                    if is_mean and calc_mean:
                        if not c_mean_table.empty:
                            c_cl_table.loc['Среднее'] = c_mean_table.loc['Среднее']

                    # Concatenate all table
                    if dk_check:
                        c_cl_table = pd.concat([c_cl_table, cl_table_dk], axis=0)
                    r_cl_table[col_var] = c_cl_table

            # Assign base row to the tables if it is not last multipunch and concatenate row table
            if is_base:
                col_table.loc['База:'] = base
                abs_table.loc['База:'] = base
            r_pivot_table[col_var] = col_table
            r_abs_table[col_var] = abs_table
            if calc_r_cl:
                if is_base:
                    c_round_table.loc['База:'] = base.round(0)
                r_round_table[col_var] = c_round_table

        # Concatenate row and set row labels
        r_pivot_table = concat_row(r_pivot_table)
        r_pivot_table = set_row_labels(r_pivot_table, meta, row_var, sub_label=True)

        r_abs_table = concat_row(r_abs_table)
        r_abs_table = set_row_labels(r_abs_table, meta, row_var, sub_label=True)

        pivot_tables.append(r_pivot_table)
        abs_tables.append(r_abs_table)

        if is_base:
            if r_pivot_table.iloc[-1].iloc[0] > client_base.iloc[0]:
                client_base = r_pivot_table.iloc[-1].copy()
        if calc_r_cl:
            r_round_table = concat_row(r_round_table)
            r_round_table = set_row_labels(r_round_table, meta, row_var, sub_label=True)
            round_tables[row_var] = r_round_table
            if is_orig:
                r_cl_table = concat_row(r_cl_table)
                r_cl_table = set_row_labels(r_cl_table, meta, row_var, sub_label=client_sublabels)
                client_tables.append(r_cl_table)
        if is_mean and calc_mean:
            r_mean_table = concat_row(r_mean_table)
            r_mean_table = set_row_labels(r_mean_table, meta, row_var, sub_label=True)
            mean_tables.append(r_mean_table)

    # Concatenate row tables to summary tables
    def concat_summary(table):
        return pd.concat(table if isinstance(table, list) else list(table.values()), axis=0).fillna(0)

    sum_pivot_table = concat_summary(pivot_tables)
    sum_abs_table = concat_summary(abs_tables)
    dummy_banner = pd.DataFrame(columns=sum_pivot_table.columns)

    sum_round_table = concat_summary(round_tables) if calc_r_cl else dummy_banner.copy()
    sum_client_table = concat_summary(client_tables) if calc_r_cl else dummy_banner.copy()
    sum_mean_table = concat_summary(mean_tables) if calc_mean and mean_tables else dummy_banner.copy()

    # Set banner columns and letters
    banner_labels = [('Total', 'Total', 'T', 't')]
    orig_ttl_value = round(df[wt_var].sum())
    client_total_round = math.ceil(orig_ttl_value / 100.0) * 100
    client_banner_labels = [('Данные', 'Данные', f'{client_total_round}')]

    if len(banner) > 1:
        client_base = client_base.iloc[1:].reset_index()
        client_base = client_base.groupby('level_0').apply(normalize_to_total,
                                                           round_ttl_base=client_total_round,
                                                           orig_ttl_base=orig_ttl_value
                                                           ).droplevel(0).sort_index().iloc[:,-1]

    # Subsample letters
    level_3_col_list = banner[1:]
    level_3_col_list = [meta.var_names_to_labels[x] for x in level_3_col_list]
    level_3_col_list = list(set(level_3_col_list))
    banner_num = len(level_3_col_list)
    banner_letters_level_3 = [chr(rep) for rep in range(65, 91) if rep != 84]
    extra_combinations_needed = max(0, banner_num - len(banner_letters_level_3))
    extra_combinations = [chr(65 + (i // 25)) + chr(65 + (i % 25)) for i in range(extra_combinations_needed)]
    banner_letters_level_3 = banner_letters_level_3[:banner_num] + extra_combinations
    banner_letters_level_3 = dict(zip(level_3_col_list, banner_letters_level_3))

    # Sub question letters
    level_4_dict = defaultdict(str)
    counter = defaultdict(int)

    # Create and set banner
    for i, pair in enumerate(sum_pivot_table.columns[1:]):
        q, v = pair
        l = meta.var_names_to_labels[q]

        index = counter[l]
        counter[l] = int(counter[l]) + 1
        # avois using lone 't' letter. still double letters are allowed
        level_4_dict[pair] = chr(97 + index) if index < 19 else chr(97 + index + 1) if index < 25 else chr(
            97 + index % 25) * (index // 25 + 1)

        value_label = meta.var_value_labels[q].get(v, str(v))
        question_label = meta.var_names_to_labels[q]

        level_3_char = banner_letters_level_3.get(meta.var_names_to_labels[q], '')
        level_4_char = level_4_dict[pair]

        banner_labels.append((question_label, value_label, level_3_char, level_4_char))
        client_banner_labels.append((question_label, value_label, client_base.iloc[i]))

    # Apply to all tables
    for table in [sum_pivot_table, sum_abs_table, sum_round_table, sum_mean_table]:
        table.columns = pd.MultiIndex.from_tuples(banner_labels, names=[None, None, None, None])
    sum_client_table.columns = pd.MultiIndex.from_tuples(client_banner_labels, names=[None, None, None])

    tables_list = [
       sum_pivot_table,
       sum_round_table,
       sum_mean_table,
       sum_client_table,
       sum_abs_table
       ]
    sheet_labels = ['Результаты', 'Округленные', 'Средние', 'Для заказчика', 'Абсолюты']

    # added z-test calculation
    if cl:
        sig_prop_tables = z_test.calc_sig_prop(sum_pivot_table, cl)
        tables_list.append(sig_prop_tables)
        sheet_labels.append(f'Результаты CL{int((1-cl)*100)}%')

        if not sum_mean_table.empty:
            sig_mean_tables = z_test.calc_sig_mean(sum_mean_table, cl)
            tables_list.append(sig_mean_tables)
            sheet_labels.append(f'Средние CL{int((1-cl)*100)}%')

    final_tables = dict(zip(sheet_labels, tables_list))

    return final_tables


def table_export(export_path, tables, empty_sheets=True, wt_var=None):
    """
    Export the tables to an Excel file with the given export path.

    Args:
        export_path (str): The path where the Excel file will be saved.
        tables (dict): A dictionary mapping sheet names to pandas DataFrames representing the tables.

    Returns:
        None
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Alignment, Font


    def index_formatter(cell):
        cell.font = Font(name='Arial', color="000000", bold=True, size=9)
        cell.fill = PatternFill(start_color="E6F0FC", end_color="E6F0FC", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


    def column_formatter(cell):
        cell.font = Font(name='Arial', color="000000", bold=True, size=9)
        cell.fill = PatternFill(start_color="E6F0FC", end_color="E6F0FC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)


    def client_index_formatter(cell):
        val = str(cell.value)
        cond_1 = val.startswith('Top') or val.startswith('Bottom')
        cond_2 = val in ['Среднее', 'Знаю', 'Не знаю']

        if cond_1 or cond_2:
            cell.font = Font(name='Arial', color="000000", bold=True, italic=True, size=9)
            cell.fill = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.font = Font(name='Arial', color="000000", bold=True, size=9)
            cell.fill = PatternFill(start_color="E6F0FC", end_color="E6F0FC", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


    def delete_empty_row(ws, idx):
        ws.delete_rows(idx)
        for mcr in ws.merged_cells:
            if idx < mcr.min_row:
                mcr.shift(row_shift=-1)
            elif idx <= mcr.max_row:
                mcr.shrink(bottom=1)

    low_sig_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    high_sig_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    low_base_fill = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")


    with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
        for sheet_name, table in tables.items():
            workbook = writer.book
            if table.empty:
                # If the table is empty, create an empty sheet in the workbook
                if empty_sheets:
                    workbook.create_sheet(title=sheet_name)
            else:
                table.to_excel(writer, sheet_name=sheet_name)

    # Reopen workbook and apply formatting
    wb = load_workbook(export_path)

    # Content format
    percentage_format = '0%'
    number_format = '0'
    decimal_format = '0.00'

    for sheet_name, table in tables.items():
        if table.empty:
            continue

        ws = wb[sheet_name]
        print(sheet_name)
        is_sig = "CL" in sheet_name
        print(is_sig)

        if sheet_name != 'Для заказчика':
            delete_empty_row(ws, 5)
            ws.merge_cells('A1:B4')
            ws.row_dimensions[3].hidden = True
            row_start = 5
        else:
            delete_empty_row(ws, 4)
            ws.merge_cells('A1:B3')
            row_start = 4

        row_intendant = 3

        for cell in ws['A']:
            index_formatter(cell)
        for i, cell in enumerate(ws['B'], start=1):
            if sheet_name != 'Для заказчика':
                index_formatter(cell)
            else:
                if cell.value == 'Среднее' and not ws.cell(row=i - 1, column=2).value.startswith('Bottom'):
                    index_formatter(cell)
                else:
                    client_index_formatter(cell)

        for cell in ws[1]:
            column_formatter(cell)
        for cell in ws[2]:
            column_formatter(cell)

        # Cell size
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.row_dimensions[1].height = 59
        ws.row_dimensions[2].height = 85

        # Data formatting
        column_len = len(table.columns)

        for i, value in enumerate(table.index, start=row_start):
            for j in range(column_len):
                col_i=j + row_intendant
                print(col_i)
                cell = ws.cell(row=i, column=col_i)

                if sheet_name == 'Абсолюты':
                    cell.number_format = number_format
                elif sheet_name != 'Средние':
                    if (value[1] == 'Среднее') and (sheet_name == 'Для заказчика'):
                        if ws.cell(row=i - 1, column=2).value.startswith('Bottom'):
                            cell.number_format = decimal_format
                        else:
                            cell.number_format = percentage_format
                    elif value[1] != 'База:':
                        cell.number_format = percentage_format
                    else:
                        cell.number_format = number_format

                    # ADD SIG COLOR
                    if is_sig and (col_i%2==1):
                        print(cell)

                        v = cell.value
                        if pd.notna(v):
                            s = str(v)
                            if "<t" in s:
                                cell.fill = low_sig_fill
                                ws.cell(row=i, column=col_i-1).fill = low_sig_fill
                            elif ">t" in s:
                                cell.fill = high_sig_fill
                                ws.cell(row=i, column=col_i-1).fill = high_sig_fill

                            elif s == "_":
                                cell.value = ""
                                cell.fill = low_base_fill
                                ws.cell(row=i, column=col_i-1).fill = low_base_fill

                else:
                    if value[1] != 'Валидные':
                        cell.number_format = decimal_format
                    else:
                        cell.number_format = number_format
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Bold total
                if j == 0:
                    cell.font = Font(name='Arial', color="000000", bold=True, size=9)
                else:
                    cell.font = Font(name='Arial', color="000000", bold=False, size=9)

        # Client table base formatting
        if sheet_name == 'Для заказчика':
            cell = ws.cell(row=1, column=1)
            cell.value = 'Выборка'
            cell.font = Font(name='Arial', color="000000", bold=True, size=9)
            cell.alignment = Alignment(horizontal='right', vertical='bottom')

        # Freeze panes
        ws.freeze_panes = f'D{row_start}'

        # Add weight spec to the A1 cell
        cell = ws.cell(row=1, column=1)
        cell.value = f"Взвешено на {wt_var}" if wt_var else 'БЕЗ ВЕСА'
        cell.font = Font(name='Arial', color="f50000", bold=True, size=12)
        cell.alignment = Alignment(horizontal='left', vertical='top')

    # Save changes
    wb.save(export_path)


def fill_empty_dimension(table, val_labels, axis):
    """
    Fill dimension in a table with missing values.
    Parameters:
        - table: The table to fill empty dimension in.
        - val_labels: A dictionary containing the full value-labels dictionary.
        - axis: The axis along which to fill the empty dimension.
                If axis is 0, the operation is performed on the rows.
                If axis is 1, the operation is performed on the columns.
    Returns:
        - The table with the empty dimension filled.
    """
    exist_codes = table.index if axis == 0 else table.columns
    missing_codes = [code
                     for code in val_labels.keys()
                     if code != 0 and code not in exist_codes]
    if axis == 0:
        table = table.reindex(table.index.union(missing_codes), fill_value=0)
    else:
        table = table.reindex(columns=table.columns.union(missing_codes), fill_value=0)

    table.sort_index(axis=axis, inplace=True)
    return table


def concat_row(table):
    """
    Concatenates the values of a dictionary of pandas DataFrames horizontally.

    Args:
        table (dict): A dictionary of pandas DataFrames.

    Returns:
        pandas.DataFrame: A new DataFrame obtained by concatenating the DataFrames
        horizontally.

    Example:
        >>> table = {'A': pd.DataFrame({'col1': [1, 2, 3], 'col2': [4, 5, 6]}),
        ...          'B': pd.DataFrame({'col1': [7, 8, 9], 'col2': [10, 11, 12]})}
        >>> concat_row(table)
           A        B
        col1 col2 col1 col2
        0    1    4    7    10
        1    2    5    8    11
        2    3    6    9    12
    """
    return pd.concat(list(table.values()), axis=1, keys=list(table.keys()))


def block_table(
        df,
        meta,
        banner,
        block,
        ttl_var='ttl',
        wt_var='wt_right',
        empty_incl=True,
        drop_mp_zeros=True,
        calc_mean=True,
):


    def calculate_block_mean_parameters(x):
        """
        Calculate the mean parameters.

        Parameters:
            x (pandas.DataFrame): The input data frame with variable and weight column.

        Returns:
            pd.Series: A series containing the calculated mean parameters:
                - 'Среднее': The weighted mean.
        """
        wt_base = x[wt_var].sum()
        base_condition = wt_base > 1

        # To avoid excess computation, use if actual base exists
        if base_condition:
            wt_mean = np.average(x.iloc[:,0], weights=x[wt_var])
        else:
            wt_mean = 0

        return pd.Series({
            'Среднее': wt_mean
        })




    def set_block_labels(table, question):
        """
        Set the row labels of a table based on the given metadata.

        Parameters:
            table (pd.DataFrame): The table whose row labels will be set.
            question (object): object with question from qre details.

        Returns:
            pd.DataFrame: The table with the updated row labels.
        """
        table = table.rename(index=question.value_labels)
        table.index = pd.MultiIndex.from_tuples(
            [(question.q_label, question.sub_label, value) for value in table.index], names=['q', 's', 'v'])
        return table


    # Initialize body variables
    body = block.vars

    # Initialize table templates
    pivot_tables = []

    # Loop through all variables and all banner variables, generating pivot tables
    for row_i, row_var in enumerate(body):
        # Don't use if empty include is set to False and variable is all nan
        empty_condition = df[row_var].isna().all()
        # If there is breaker inside the MP or gris block it is possible to duplicate columns

        # # какой-то воркэраундр нужен
        # if empty_condition:
        #     pass

        # Analyze the variable type
        is_mean = row_var in meta.scale_vars

        rv_type = meta.q_type[row_var]
        # If base row should be applied

        # Initialize row variable tables
        r_pivot_table = {}


        # Iterate through each banner variable, creating table and storing into row table
        for col_var in banner:
            # If banner var is empty, workaround
            banner_empty_condition = df[col_var].isna().all()
            if banner_empty_condition:
                pass

            # Calculate pivot table, fill nan with 0
            col_table = df.groupby([row_var, col_var])[wt_var].sum().unstack().fillna(0)

            # Drop zeroes from column
            if 0 in col_table.columns:
                col_table.drop(0, axis=1, inplace=True)

            # Fill rows with missed values for total and fill columns for the rest of banner
            if empty_incl:
                if col_var == ttl_var:
                    col_table = fill_empty_dimension(
                        table=col_table,
                        val_labels=meta.var_value_labels[row_var],
                        axis=0
                    )
                elif col_var == 'fltr_ttl':
                    pass
                else:
                    col_table = fill_empty_dimension(
                        table=col_table,
                        val_labels=meta.var_value_labels[col_var],
                        axis=1
                    )

            # Calculate base for table
            base = col_table.sum()

            # Convert to percentage
            col_table = col_table / base

            # Drop zeroes from rows for multipunch
            if all([drop_mp_zeros, rv_type == 'multipunch', 0 in col_table.index]):
                col_table.drop(0, axis=0, inplace=True)

            # Calculate mean tables
            if all([calc_mean, is_mean]):
                temp_mean_series = df.loc[df[col_var] != 0, [row_var, col_var, wt_var]]
                if row_var != col_var:
                    temp_mean_series[row_var].replace(meta.scale_recode_for_mean[row_var], inplace=True)
                # Handler to cross the variable to itself
                else:
                    temp_mean_series.columns = ['dummy_row', col_var, wt_var]
                    temp_mean_series['dummy_row'].replace(meta.scale_recode_for_mean[row_var], inplace=True)

                temp_mean_series.dropna(axis=0, inplace=True)
                if temp_mean_series.empty:
                    c_mean_table = pd.DataFrame(
                        data={1: [np.nan]},
                        index=['Среднее']
                    ).rename_axis(mapper=col_var, axis='columns')
                else:
                    c_mean_table = temp_mean_series.groupby(col_var, dropna=False).apply(
                        calculate_block_mean_parameters).T

                # Fill empty if necessary
                if col_var != ttl_var:
                    c_mean_table = fill_empty_dimension(
                        table=c_mean_table,
                        val_labels=meta.var_value_labels[col_var],
                        axis=1
                    )

                # Append to the main table after scale codes
                scale_ind = len([v for v in
                             list(meta.scale_recode_for_mean[row_var].values()) if pd.notna(v)])
                col_table = pd.concat([
                    col_table.iloc[:scale_ind],
                    c_mean_table,
                    col_table.iloc[scale_ind:]
                ], axis=0)

            # Set row table to the dict
            r_pivot_table[col_var] = col_table

        # Concatenate row and set row labels
        r_pivot_table = concat_row(r_pivot_table)
        r_pivot_table = set_block_labels(r_pivot_table, block.questions[row_i])

        pivot_tables.append(r_pivot_table)

    # Concatenate row tables to summary tables
    sum_pivot_table = pd.concat(pivot_tables).fillna(0)

    return sum_pivot_table